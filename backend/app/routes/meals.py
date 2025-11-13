
from fastapi import APIRouter, HTTPException, Header, Depends, Query
from app.utils.nutrition import calculate_nutrition_targets
from app.core.supabase import verify_token, supabase
from app.models.user import UserCreate
from postgrest.exceptions import APIError

from typing import Optional
from datetime import datetime, date as date_cls, time as time_cls, timedelta, timezone
from zoneinfo import ZoneInfo

router = APIRouter()

def get_current_user_id(authorization: str = Header(...)) -> str:
    user_id = verify_token(authorization)
    return user_id


@router.get("/history_meals")
def get_meal_history(user_id: str = Depends(get_current_user_id)):
    try:
        history = supabase.table("meals").select("*").eq("user_id", user_id).order("date_creation", desc=True).execute()
        return history.data
    except APIError as e:
        raise HTTPException(status_code=500, detail=str(e))
    
    
@router.get("/history_meals/{meal_id}")
def get_meal_detail(meal_id: str, user_id: str = Depends(get_current_user_id)):
    try:
        meal = supabase.table("meals").select("*").eq("id", meal_id).eq("user_id", user_id).execute()
        meal_items = supabase.table("meal_items").select("*").eq("meal_id", meal_id).execute()
        if not meal.data:
            raise HTTPException(status_code=404, detail="Meal not found")
        return {"meal": meal.data[0], "items": meal_items.data}
    except APIError as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/delete_meal/{meal_id}")
def delete_meal(meal_id: str, user_id: str = Depends(get_current_user_id)):
    try:
        meal = supabase.table("meals").select("*").eq("id", meal_id).eq("user_id", user_id).execute()
        if not meal.data:
            raise HTTPException(status_code=404, detail="Meal not found")
        supabase.table("meal_items").delete().eq("meal_id", meal_id).execute()
        supabase.table("meals").delete().eq("id", meal_id).eq("user_id", user_id).execute()
        return {"detail": "Meal deleted successfully"}
    except APIError as e:
        raise HTTPException(status_code=500, detail=str(e))



from typing import Optional, Tuple
from datetime import datetime, date as date_cls, time as time_cls, timedelta, timezone

try:
    from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
except Exception:
    ZoneInfo = None
    ZoneInfoNotFoundError = Exception

def resolve_tz(tz_name: str):
    """
    Devuelve tzinfo robusto:
    - Intenta ZoneInfo(tz_name)
    - Si falla y es Lima: -05:00 fijo (Perú no usa DST)
    - Si falla cualquier otra: UTC
    """
    if ZoneInfo is not None:
        try:
            return ZoneInfo(tz_name)
        except ZoneInfoNotFoundError:
            pass
    if tz_name == "America/Lima":
        return timezone(timedelta(hours=-5))
    return timezone.utc

def _day_range_utc(date_str: Optional[str], tz_name: str = "America/Lima") -> tuple[str, str, str]:
    """
    Convierte una fecha local (YYYY-MM-DD) a rango UTC [start, end) ISO8601.
    Si no se pasa fecha, usa la fecha “hoy” en la TZ indicada.
    """
    tz = resolve_tz(tz_name)

    if date_str:
        target_date = date_cls.fromisoformat(date_str)
    else:
        target_date = datetime.now(tz).date()

    start_local = datetime.combine(target_date, time_cls.min).replace(tzinfo=tz)
    end_local   = start_local + timedelta(days=1)

    start_utc = start_local.astimezone(timezone.utc).isoformat()
    end_utc   = end_local.astimezone(timezone.utc).isoformat()
    return (target_date.isoformat(), start_utc, end_utc)

import io
import csv
from openpyxl import Workbook

@router.get("/meals/day")
def get_meals_and_summary_for_day(
    date: Optional[str] = Query(
        default=None,
        description="Fecha en formato YYYY-MM-DD. Por defecto, la fecha actual en America/Lima."
    ),
    tz: str = Query(
        default="America/Lima",
        description="Timezone IANA para calcular el día local (ej. America/Lima)."
    ),
    user_id: str = Depends(get_current_user_id),
):
    """
    Devuelve:
    - date: fecha local solicitada
    - timezone: timezone usado
    - targets: objetivos diarios del usuario (required_* y metadatos)
    - totals: sumatoria consumida en el día (cal, prot, carb, fat)
    - meals_count: número de comidas del día
    - meals: lista de comidas del día (para listar/depurar/thumbnail)
    """
    try:
        # Rango del día en UTC (robusto vs date(date_creation) = ...)
        local_date, start_utc, end_utc = _day_range_utc(date, tz)

        # 1) Traer comidas del día del usuario
        meals_res = (
            supabase.table("meals")
            .select(
                "id,user_id,date_creation,img_url,recommendation,"
                "total_calories,total_carbs_g,total_fat_g,total_protein_g"
            )
            .eq("user_id", user_id)
            .gte("date_creation", start_utc)
            .lt("date_creation", end_utc)
            .order("date_creation", desc=False)
            .execute()
        )
        meals = meals_res.data or []

        # 2) Sumar totales
        def f(x):  # cast seguro a float
            try:
                return float(x) if x is not None else 0.0
            except Exception:
                return 0.0

        totals = {
            "calories": sum(f(m.get("total_calories")) for m in meals),
            "protein_g": sum(f(m.get("total_protein_g")) for m in meals),
            "carbs_g":   sum(f(m.get("total_carbs_g")) for m in meals),
            "fat_g":     sum(f(m.get("total_fat_g")) for m in meals),
        }

        # 3) Traer objetivos del usuario para las barras (y header “bonito”)
        #    Usamos columnas ya calculadas si existen en tu tabla users.
        user_res = (
            supabase.table("users")
            .select(
                "required_calories,required_protein_g,required_fat_g,required_carbs_g,"
                "objective_id,activity_level_id"
            )
            .eq("id", user_id)
            .execute()
        )
        user_row = (user_res.data or [{}])[0]

        targets = {
            "required_calories": user_row.get("required_calories"),
            "required_protein_g": user_row.get("required_protein_g"),
            "required_fat_g": user_row.get("required_fat_g"),
            "required_carbs_g": user_row.get("required_carbs_g"),
            "objective": user_row.get("objective_id"),
            "activity_level": user_row.get("activity_level_id"),
        }

        return {
            "date": local_date,
            "timezone": tz,
            "targets": targets,
            "totals": totals,
            "meals_count": len(meals),
            "meals": meals,
        }

    except APIError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except ValueError:
        # fecha malformateada
        raise HTTPException(
            status_code=400,
            detail="El parámetro 'date' debe tener formato YYYY-MM-DD."
        )
        

def _range_utc(
    from_date: Optional[str],
    to_date: Optional[str],
    tz_name: str,
) -> Optional[tuple[str, str]]:
    """
    Convierte un rango local [from_date, to_date] a rango UTC [start, end).

    - from_date/to_date: YYYY-MM-DD
    - Si ambos son None: devuelve None → se interpreta como "todo".
    - Si solo from_date: to_date = from_date.

    Devuelve (start_utc_iso, end_utc_iso).
    """
    if not from_date and not to_date:
        return None  # sin filtro por fecha

    tz = resolve_tz(tz_name)

    if not from_date and to_date:
        # si solo llega to_date, usamos ese día como único día
        from_date = to_date

    # Asegurar a los type checkers que from_date ya no es None
    assert from_date is not None

    start_date = date_cls.fromisoformat(from_date)
    end_date = date_cls.fromisoformat(to_date) if to_date else start_date

    if end_date < start_date:
        raise ValueError("to_date no puede ser anterior a from_date")

    start_local = datetime.combine(start_date, time_cls.min).replace(tzinfo=tz)
    end_local = datetime.combine(end_date + timedelta(days=1), time_cls.min).replace(tzinfo=tz)

    start_utc = start_local.astimezone(timezone.utc).isoformat()
    end_utc = end_local.astimezone(timezone.utc).isoformat()
    return start_utc, end_utc


def _local_date_time(iso_str: str, tz_name: str) -> tuple[str, str]:
    """
    Convierte date_creation ISO (UTC u offset) a (fecha_local, hora_local) en tz_name.
    """
    tz = resolve_tz(tz_name)
    dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
    dt_local = dt.astimezone(tz)
    return dt_local.date().isoformat(), dt_local.strftime("%H:%M")


@router.get("/meals/export_history")
def export_meals_history(
    from_date: Optional[str] = Query(
        default=None,
        description="Fecha inicio (YYYY-MM-DD). Si se omite junto con to_date, se exporta todo."
    ),
    to_date: Optional[str] = Query(
        default=None,
        description="Fecha fin (YYYY-MM-DD). Si solo se pasa from_date, se usa el mismo día."
    ),
    format: str = Query(
        default="xlsx",
        regex="^(csv|xlsx)$",
        description="Formato de exportación: csv o xlsx"
    ),
    tz: str = Query(
        default="America/Lima",
        description="Timezone IANA para mostrar fecha/hora (ej. America/Lima)."
    ),
    user_id: str = Depends(get_current_user_id),
):
    """
    Exporta el historial de comidas del usuario:

    - Si NO se envían from_date/to_date -> TODO el historial.
    - Si se envía from_date (y opcional to_date) -> comidas solo en ese rango local.
    """
    try:
        # 1) Construir query base
        query = (
            supabase.table("meals")
            .select(
                "id,date_creation,img_url,"
                "total_calories,total_protein_g,total_carbs_g,total_fat_g"
            )
            .eq("user_id", user_id)
        )

        # 2) Aplicar rango si corresponde
        range_utc = _range_utc(from_date, to_date, tz)
        if range_utc is not None:
            start_utc, end_utc = range_utc
            query = query.gte("date_creation", start_utc).lt("date_creation", end_utc)

        # 3) Ejecutar query
        meals_res = query.order("date_creation", desc=False).execute()
        meals = meals_res.data or []

        # 4) Metadata para el archivo
        if from_date or to_date:
            rango_txt = f"{from_date or to_date} → {to_date or from_date}"
            filename = f"nutriapp_meals_{from_date or to_date}_to_{to_date or from_date}.{format}"
        else:
            rango_txt = "Todo el historial"
            filename = f"nutriapp_meals_history.{format}"

        # 5) CSV
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)

            writer.writerow([
                "Fecha",
                "Hora",
                "ID comida",
                "Calorías",
                "Proteínas (g)",
                "Carbohidratos (g)",
                "Grasas (g)",
            ])

            for m in meals:
                fecha_local, hora_local = _local_date_time(m["date_creation"], tz)
                writer.writerow([
                    fecha_local,
                    hora_local,
                    m["id"],
                    m.get("total_calories", 0),
                    m.get("total_protein_g", 0),
                    m.get("total_carbs_g", 0),
                    m.get("total_fat_g", 0),
                ])

            csv_data = output.getvalue()
            return Response(
                content=csv_data,
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                },
            )

        # 6) XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial comidas"

        ws["A1"] = "Historial de comidas"
        ws["A2"] = f"Rango: {rango_txt}"
        ws["A3"] = f"Zona horaria: {tz}"
        ws["A4"] = f"Número de registros: {len(meals)}"

        header_row = 6
        headers = [
            "Fecha",
            "Hora",
            "ID comida",
            "Calorías",
            "Proteínas (g)",
            "Carbohidratos (g)",
            "Grasas (g)",
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=h)

        row = header_row + 1
        for m in meals:
            fecha_local, hora_local = _local_date_time(m["date_creation"], tz)
            ws.cell(row=row, column=1, value=fecha_local)
            ws.cell(row=row, column=2, value=hora_local)
            ws.cell(row=row, column=3, value=m["id"])
            ws.cell(row=row, column=4, value=float(m.get("total_calories") or 0))
            ws.cell(row=row, column=5, value=float(m.get("total_protein_g") or 0))
            ws.cell(row=row, column=6, value=float(m.get("total_carbs_g") or 0))
            ws.cell(row=row, column=7, value=float(m.get("total_fat_g") or 0))
            row += 1

        # Totales
        ws.cell(row=row, column=2, value="Totales")
        ws.cell(row=row, column=4, value=f'=SUM(D{header_row+1}:D{row-1})')
        ws.cell(row=row, column=5, value=f'=SUM(E{header_row+1}:E{row-1})')
        ws.cell(row=row, column=6, value=f'=SUM(F{header_row+1}:F{row-1})')
        ws.cell(row=row, column=7, value=f'=SUM(G{header_row+1}:G{row-1})')

        # Ancho columnas
        col_widths = [12, 8, 30, 12, 14, 16, 12]
        for i, w in enumerate(col_widths, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = w

        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()

        return Response(
            content=xlsx_data,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )

    except ValueError as e:
        # from_date/to_date mal formateadas o rango inválido
        raise HTTPException(status_code=400, detail=str(e))
    except APIError as e:
        raise HTTPException(status_code=500, detail=str(e))